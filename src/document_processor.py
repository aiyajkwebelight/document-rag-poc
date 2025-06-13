from litellm import embedding
import os
from dotenv import load_dotenv
from langchain_text_splitters import RecursiveCharacterTextSplitter
from qdrant_client import QdrantClient
from qdrant_client.http.models import Distance, VectorParams

import PyPDF2
import csv
from openpyxl import load_workbook

load_dotenv()

UPLOAD_FOLDER = 'files'
ALLOWED_EXTENSIONS = {'txt', 'pdf', 'csv', 'xlsx'}
CHUNK_SIZE = 1000  
CHUNK_OVERLAP = 100  

QDRANT_PATH = "qdrant_data"
QDRANT_COLLECTION = "document_embeddings"
QDRANT_VECTOR_SIZE = 768  # Should match your embedding size

# qdrant_client = QdrantClient(path=QDRANT_PATH, prefer_grpc=False)
qdrant_client = QdrantClient(host="localhost", port=6333)
try:
    qdrant_client.create_collection(
        collection_name=QDRANT_COLLECTION,
        vectors_config=VectorParams(size=QDRANT_VECTOR_SIZE, distance=Distance.COSINE),
    )
except Exception:
    pass  # Collection may already exist



def extract_text_from_file(file_path):
    """Extract text content from a file based on its extension"""
    try:
        file_extension = file_path.rsplit('.', 1)[1].lower()
        
        if file_extension == 'txt':
            try:
                with open(file_path, 'r', encoding='utf-8', errors='replace') as file:
                    return file.read()
            except Exception as e:
                print(f"Error reading txt file: {e}")
                return ""
                
        elif file_extension == 'pdf':
            try:
                text = ""
                with open(file_path, 'rb') as file: 
                    pdf_reader = PyPDF2.PdfReader(file)
                    for page_num in range(len(pdf_reader.pages)):
                        text += pdf_reader.pages[page_num].extract_text() + "\n"
                return text
            except Exception as e:
                print(f"Error reading PDF file: {e}")
                return ""
            
        elif file_extension == 'csv':
            try:
                text = ""
                with open(file_path, 'r', encoding='utf-8', errors='replace') as file:
                    reader = csv.reader(file)
                    for row in reader:
                        text += ", ".join(row) + "\n"
                return text
            except Exception as e:
                print(f"Error reading CSV file: {e}")
                return ""
            
        elif file_extension == 'xlsx':
            try:
                text = ""
                workbook = load_workbook(file_path, read_only=True)
                for sheet in workbook:
                    for row in sheet.iter_rows(values_only=True):
                        text += ", ".join([str(cell) if cell is not None else "" for cell in row]) + "\n"
                return text
            except Exception as e:
                print(f"Error reading XLSX file: {e}")
                return ""
        
        return ""
    except Exception as e:
        print(f"Error extracting text from file: {e}")
        return ""

def create_chunks(text, chunk_size=CHUNK_SIZE, overlap=CHUNK_OVERLAP):
    """Split text into chunks with overlap using LangChain's RecursiveCharacterTextSplitter
    
    This function replaces the custom chunking implementation with LangChain's more sophisticated
    text splitting approach that respects semantic boundaries (paragraphs, sentences, etc.)
    
    Args:
        text (str): The text to split into chunks
        chunk_size (int): Maximum size of each chunk
        overlap (int): Number of characters to overlap between chunks
        
    Returns:
        list: List of text chunks
    """
    # Return empty list for empty text
    if not text:
        return []
    
    try:    
        # Create the text splitter with the specified configuration
        text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=chunk_size,
            chunk_overlap=overlap,
            length_function=len,
        )
        
        # Split the text into chunks
        chunks = text_splitter.split_text(text)
        
        return chunks
    except Exception as e:
        print(f"Error in create_chunks: {e}")
        # Fallback to a simple chunking method if LangChain fails
        chunks = []
        start = 0
        text_length = len(text)
        
        while start < text_length:
            end = min(start + chunk_size, text_length)
            chunks.append(text[start:end])
            start = end - overlap if end < text_length else text_length
        
        return chunks

def generate_embedding(text):
    """Generate embedding vector for given text"""
    try:
        response = embedding(
            input=[text],
            model=os.getenv("EMBEDDING_MODEL"),
            api_key=os.getenv("GEMINI_API_KEY"),
        )
        return response['data'][0]['embedding']
    except Exception as e:
        print(f"Error generating embedding: {e}")
        return [0.0] * 768  

def search_documents(query_text, num_results=5):
    """
    Search for documents matching the query and return simplified results
    
    Args:
        query_text (str): The text to search for
        num_results (int): Number of results to return
        
    Returns:
        list: List of dictionaries containing document content, filename, and chunk index or id
    """
    try:
        query_embedding = generate_embedding(query_text)
        search_result = qdrant_client.search(
            collection_name=QDRANT_COLLECTION,
            query_vector=query_embedding,
            limit=num_results,
            with_payload=True
        )
        simplified_results = []
        for hit in search_result:
            payload = hit.payload or {}
            source_type = payload.get("source", "document")
            simplified_results.append({
                "document": payload.get("document", ""),
                "filename": payload.get("filename", "Unknown"),
                "id": payload.get("id", 0),
                "source_type": source_type
            })
        return simplified_results
    except Exception as e:
        print(f"Error searching documents: {e}")
        return []